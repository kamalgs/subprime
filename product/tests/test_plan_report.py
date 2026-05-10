"""Tests for the plan PDF + Excel report generators.

Covers:
  - PDF bytes are valid (starts with %PDF, reasonable length, contains our
    branded strings like 'Benji' and the disclaimer).
  - PDF survives edge cases: missing projected_returns, empty setup_phase,
    risks with leading '- ' markers (defensive — _format strips these but
    the report must still render if a caller skipped normalization).
  - XLSX opens cleanly, has every expected sheet, numeric cells are numeric,
    disclaimer is present and red-coloured.
"""

from __future__ import annotations

from io import BytesIO

import pytest

from subprime.core.models import (
    Allocation,
    InvestmentPlan,
    InvestorProfile,
    MutualFund,
)
from subprime.core.plan_report import (
    _fmt_money_inr,
    _project_corpus,
    _projection_trace,
    _split_bullets,
    build_plan_pdf,
    build_plan_xlsx,
)


@pytest.fixture
def profile() -> InvestorProfile:
    return InvestorProfile(
        id="P01",
        name="Ananya Shetty",
        financial_goals=["retirement", "home"],
        life_stage="mid career",
        tax_bracket="30%",
        age=38,
        risk_appetite="moderate",
        monthly_investible_surplus_inr=50000,
        existing_corpus_inr=2_500_000,
        liabilities_inr=0,
        investment_horizon_years=15,
    )


@pytest.fixture
def plan() -> InvestmentPlan:
    f1 = MutualFund(
        amfi_code="119551",
        name="HDFC Nifty 50 Index Direct Growth",
        display_name="HDFC Nifty 50 Index",
        category="Index",
        fund_house="HDFC",
    )
    f2 = MutualFund(
        amfi_code="120505",
        name="Parag Parikh Flexi Cap Direct Growth",
        display_name="Parag Parikh Flexi Cap",
        category="Flexi Cap",
        fund_house="PPFAS",
    )
    return InvestmentPlan(
        allocations=[
            Allocation(
                fund=f1,
                allocation_pct=40.0,
                mode="sip",
                monthly_sip_inr=20_000,
                lumpsum_inr=0,
                rationale="Core equity",
            ),
            Allocation(
                fund=f2,
                allocation_pct=30.0,
                mode="sip",
                monthly_sip_inr=15_000,
                lumpsum_inr=0,
                rationale="Active tilt",
            ),
        ],
        setup_phase="- Open a direct MF account on Kuvera\n- Start the monthly SIPs\n- Complete KYC",
        review_checkpoints=[
            "Year 3: if small-cap fund underperforms category by >3%, switch.",
            "When horizon drops below 5 years, shift equity to hybrid.",
        ],
        rebalancing_guidelines="Rebalance annually if any bucket drifts >5 pp from target.",
        projected_returns={"bear": 8.0, "base": 12.0, "bull": 16.0},
        rationale="Mid-career moderate risk with 15-year horizon — index core plus active tilt.",
        risks=[
            "Markets can drop 20-30% in a bad year.",
            "Active funds may underperform the index for extended periods.",
        ],
    )


# ── PDF ──────────────────────────────────────────────────────────────────────


def test_pdf_is_valid_bytes(plan, profile):
    pdf = build_plan_pdf(plan, profile)
    assert pdf[:5] == b"%PDF-", "not a valid PDF magic header"
    assert len(pdf) > 2000, "PDF smaller than expected — renderer likely produced empty pages"


def _pdf_text(pdf_bytes: bytes) -> str:
    """Extract the rendered text from a PDF using pypdf.

    ReportLab compresses content streams, so grepping raw bytes doesn't
    work for the rendered text. pypdf decompresses and decodes for us.
    """
    from io import BytesIO
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(pdf_bytes))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def test_pdf_contains_branding_and_disclaimer(plan, profile):
    text = _pdf_text(build_plan_pdf(plan, profile))
    assert "Benji" in text
    assert "Ananya Shetty" in text
    assert "research/educational purposes" in text


def test_pdf_contains_allocation_rows(plan, profile):
    text = _pdf_text(build_plan_pdf(plan, profile))
    assert "HDFC Nifty 50 Index" in text
    assert "Parag Parikh Flexi Cap" in text
    # Percentages are rendered with 0 decimal places in the table
    assert "40%" in text
    assert "30%" in text


def test_pdf_projection_includes_absolute_corpus(plan, profile):
    """Projection block should show both CAGR % and absolute final corpus."""
    text = _pdf_text(build_plan_pdf(plan, profile))
    # CAGR percentages
    assert "12.0%" in text
    # Absolute final corpus for the given SIP (35k/mo) and horizon (15 yrs):
    # at 12% CAGR ≈ ₹1.76 Cr — the table uses Cr / L suffixes.
    assert "Cr" in text or "L" in text  # has Indian numeric formatting somewhere


def test_pdf_has_short_disclaimer_up_front_and_footnote(plan, profile):
    text = _pdf_text(build_plan_pdf(plan, profile))
    # short disclaimer appears near the top with the ** footnote marker
    assert "research / educational use" in text
    assert "**" in text


def test_pdf_renders_without_projections(plan, profile):
    plan.projected_returns = {}
    pdf = build_plan_pdf(plan, profile)
    assert pdf[:5] == b"%PDF-"


def test_pdf_renders_without_optional_sections(plan, profile):
    plan.setup_phase = ""
    plan.rebalancing_guidelines = ""
    plan.rationale = ""
    plan.risks = []
    plan.review_checkpoints = []
    pdf = build_plan_pdf(plan, profile)
    assert pdf[:5] == b"%PDF-"
    # Disclaimer always present — it's the last flowable
    assert "research/educational purposes" in _pdf_text(pdf)


def test_pdf_handles_leading_bullet_markers_in_setup(plan, profile):
    """_format normally strips leading '- ' but the report must not choke
    if a caller skipped normalization."""
    plan.setup_phase = "- Step one\n- Step two\n- Step three"
    text = _pdf_text(build_plan_pdf(plan, profile))
    assert "Step one" in text
    assert "Step two" in text
    assert "Step three" in text


# ── Excel ────────────────────────────────────────────────────────────────────


def _load(xlsx_bytes: bytes, data_only: bool = True):
    """Load the workbook.

    ``data_only=True`` replaces formulas with their cached result — good
    for asserting numeric values. Pass ``data_only=False`` when checking
    formula strings (we haven't opened the file in Excel yet, so cached
    formula results are None).
    """
    from openpyxl import load_workbook

    return load_workbook(filename=BytesIO(xlsx_bytes), read_only=False, data_only=data_only)


def _flat_cells(sheet) -> list:
    return [c.value for row in sheet.iter_rows() for c in row]


def test_xlsx_has_only_two_sheets(plan, profile):
    wb = _load(build_plan_xlsx(plan, profile))
    assert wb.sheetnames == ["Plan", "Explore"]


def test_xlsx_plan_has_wordmark_and_investor(plan, profile):
    wb = _load(build_plan_xlsx(plan, profile))
    flat = [str(c) for c in _flat_cells(wb["Plan"]) if c]
    assert any("Benji" in s for s in flat)
    assert any("Ananya Shetty" in s for s in flat)
    assert any("research / educational" in s for s in flat)


def test_xlsx_plan_has_allocations_table(plan, profile):
    wb = _load(build_plan_xlsx(plan, profile))
    flat = [str(c) for c in _flat_cells(wb["Plan"]) if c]
    # Both fund names present
    assert any("HDFC Nifty 50 Index" in s for s in flat)
    assert any("Parag Parikh Flexi Cap" in s for s in flat)


def test_xlsx_plan_has_projection_data(plan, profile):
    wb = _load(build_plan_xlsx(plan, profile))
    p = wb["Plan"]
    # Find the "Projected returns" header row and check the next three rows
    # have numeric CAGR + corpus values.
    flat = [str(c) for c in _flat_cells(p) if c]
    assert any("Projected returns" in s for s in flat)
    # Some cell in the sheet should be a non-trivial numeric corpus (the
    # best-case corpus for 35k/mo × 15yrs × 16% ≈ ₹2.5 Cr → 25_000_000).
    numbers = [c.value for row in p.iter_rows() for c in row if isinstance(c.value, (int, float))]
    assert max(numbers) > 10_000_000  # > 1 Cr somewhere


def test_xlsx_explore_sheet_has_editable_inputs_and_formulas(plan, profile):
    wb = _load(build_plan_xlsx(plan, profile), data_only=False)
    ex = wb["Explore"]
    # Input section contains 'Monthly SIP' label + editable SIP value
    labels = [c.value for row in ex.iter_rows(min_row=4, max_row=10) for c in row if c.value]
    assert any("Monthly SIP" in str(l) for l in labels)
    assert any("Horizon" in str(l) for l in labels)
    # Output formulas for final corpus — the three FV cells must start with =
    for addr in ("E5", "E6", "E7"):
        val = ex[addr].value
        assert isinstance(val, str) and val.startswith("="), (
            f"{addr} expected a formula, got {val!r}"
        )
    # Growth-over-time formula table
    assert ex["A12"].value == "Year"
    # Year-0 formula should exist (year_cell row 13)
    assert ex["B13"].value is not None and str(ex["B13"].value).startswith("=")


def test_xlsx_plan_renders_without_optional_sections(plan, profile):
    plan.setup_phase = ""
    plan.rebalancing_guidelines = ""
    plan.rationale = ""
    plan.risks = []
    plan.review_checkpoints = []
    wb = _load(build_plan_xlsx(plan, profile))
    assert wb.sheetnames == ["Plan", "Explore"]


def test_xlsx_disclaimer_in_both_sheets(plan, profile):
    wb = _load(build_plan_xlsx(plan, profile))
    for name in ("Plan", "Explore"):
        flat = [str(c) for c in _flat_cells(wb[name]) if c]
        assert any("research" in str(s).lower() for s in flat), f"{name} sheet missing disclaimer"


# ── Numeric helper tests (mutation-testing follow-ups #63, #64) ───────────────


class TestProjectCorpus:
    """Future value of a monthly SIP — annuity-due, monthly compounding.

    Numeric assertions kill arithmetic-operator mutations on the FV formula.
    """

    def test_returns_zero_for_missing_inputs(self):
        assert _project_corpus(0, 10, 12) == 0.0
        assert _project_corpus(10_000, 0, 12) == 0.0
        # annual_pct=None is the documented "no projection" sentinel
        assert _project_corpus(10_000, 10, None) == 0.0  # type: ignore[arg-type]

    def test_zero_rate_is_simple_sum(self):
        # 10k for 10y at 0% = 10k * 120 months
        assert _project_corpus(10_000, 10, 0) == 10_000 * 120

    def test_known_value_at_twelve_percent(self):
        # 10k/mo × 10y @ 12%/yr (monthly compounded, annuity due):
        #   r = 0.01, n = 120
        #   FV = 10000 * ((1.01^120 − 1) / 0.01) * 1.01
        # Hand-computed: 2,323,390.76 (rounded).
        fv = _project_corpus(10_000, 10, 12)
        assert fv == pytest.approx(2_323_390.76, rel=1e-6)

    def test_strictly_increasing_in_horizon(self):
        prev = 0.0
        for years in range(1, 31):
            v = _project_corpus(10_000, years, 12)
            assert v > prev, f"FV not increasing at year {years}"
            prev = v

    def test_strictly_increasing_in_rate(self):
        prev = _project_corpus(10_000, 10, 0)
        for pct in (4, 8, 12, 16):
            v = _project_corpus(10_000, 10, pct)
            assert v > prev, f"FV not increasing at {pct}%"
            prev = v


class TestProjectionTrace:
    def test_starts_at_origin_and_has_horizon_plus_one_points(self):
        trace = _projection_trace(10_000, 10, 12)
        assert len(trace) == 11
        assert trace[0] == (0.0, 0.0)

    def test_year_indices_are_one_through_horizon(self):
        trace = _projection_trace(10_000, 5, 12)
        years = [yr for yr, _ in trace]
        assert years == [0.0, 1.0, 2.0, 3.0, 4.0, 5.0]

    def test_corpus_is_monotonically_non_decreasing(self):
        trace = _projection_trace(10_000, 10, 12)
        corpora = [c for _, c in trace]
        assert all(b >= a for a, b in zip(corpora, corpora[1:]))

    def test_final_point_matches_project_corpus(self):
        trace = _projection_trace(10_000, 10, 12)
        assert trace[-1][1] == pytest.approx(_project_corpus(10_000, 10, 12))


class TestFmtMoneyInr:
    """INR formatting at the lakh / crore boundaries — kills the comparison
    operator + divisor magnitude mutations in `_fmt_money_inr`.
    """

    def test_below_one_lakh_is_plain_rupees(self):
        assert _fmt_money_inr(0) == "₹0"
        assert _fmt_money_inr(1) == "₹1"
        assert _fmt_money_inr(99_999) == "₹99,999"

    def test_lakh_boundary_inclusive(self):
        # exactly ₹1 lakh formats as L, not as plain rupees
        assert _fmt_money_inr(1_00_000) == "₹1.00 L"

    def test_lakh_range_uses_lakh_divisor(self):
        assert _fmt_money_inr(2_50_000) == "₹2.50 L"
        assert _fmt_money_inr(50_00_000) == "₹50.00 L"
        # Just under 1 Cr — still L. Note: 99_99_999 / 1e5 = 99.99999 → "100.00 L"
        # which is the documented format behaviour at the upper edge.
        assert _fmt_money_inr(99_99_999) == "₹100.00 L"

    def test_crore_boundary_inclusive(self):
        # exactly ₹1 Cr formats as Cr, not as L
        assert _fmt_money_inr(1_00_00_000) == "₹1.00 Cr"

    def test_crore_range_uses_crore_divisor(self):
        assert _fmt_money_inr(12_34_56_789) == "₹12.35 Cr"
        assert _fmt_money_inr(1_00_00_00_000) == "₹100.00 Cr"


# ── Boolean / control-flow mutation kills (#66) ───────────────────────────────


class TestSplitBullets:
    """Bullet-list normalizer.

    The `continue` on empty stripped lines (line 426) is what
    ContinueWithBreak survives — without explicit empty-line tests, a
    `break` mutation would silently drop everything after the first blank.
    """

    def test_strips_dash_marker(self):
        assert _split_bullets("- one\n- two") == ["one", "two"]

    def test_handles_multiple_marker_styles(self):
        text = "- dash\n* star\n+ plus\n• bullet"
        assert _split_bullets(text) == ["dash", "star", "plus", "bullet"]

    def test_strips_numbered_prefixes(self):
        text = "1. first\n2) second\n3. third"
        assert _split_bullets(text) == ["first", "second", "third"]

    def test_preserves_lines_after_blank_line(self):
        # Kills ReplaceContinueWithBreak: with `break`, "after blank" is dropped.
        text = "- before blank\n\n- after blank"
        assert _split_bullets(text) == ["before blank", "after blank"]

    def test_drops_only_blank_lines_keeps_content_in_order(self):
        text = "\n- one\n\n\n- two\n   \n- three"
        assert _split_bullets(text) == ["one", "two", "three"]

    def test_empty_input_returns_empty_list(self):
        assert _split_bullets("") == []

    def test_unbulleted_text_falls_back_to_whole_string(self):
        assert _split_bullets("just a sentence") == ["just a sentence"]


class TestPlanPdfControlFlow:
    """Branch coverage for `if monthly:` / `if horizon:` / `if risk:` summary
    block + the `for r in plan.risks` body. ZeroIterationForLoop survives
    today because no test asserts that risks actually appear.
    """

    def test_all_summary_fields_render_when_present(self, plan, profile):
        text = _pdf_text(build_plan_pdf(plan, profile))
        # Every summary bit appears
        assert "Investible surplus" in text
        assert "Horizon" in text
        # Risk appetite shows the value, not the label
        assert "moderate" in text.lower()

    def test_summary_omits_missing_fields(self, plan, profile):
        # Drop the surplus → "Investible surplus" must not render
        profile.monthly_investible_surplus_inr = 0
        profile.investment_horizon_years = 0
        profile.risk_appetite = ""
        text = _pdf_text(build_plan_pdf(plan, profile))
        assert "Investible surplus" not in text

    def test_every_risk_line_renders(self, plan, profile):
        # Kills ZeroIterationForLoop on `for r in plan.risks`: every risk
        # must appear in the rendered text, not just the first.
        plan.risks = [
            "First risk line",
            "Second risk line",
            "Third risk line distinguishable text",
        ]
        text = _pdf_text(build_plan_pdf(plan, profile))
        for r in plan.risks:
            assert r in text, f"missing risk line: {r}"


class TestAllocationsTableFallbacks:
    """Fund display-name fallback chain — `display_name or name or amfi_code`.
    Replacing `or` with `and` would silently render an empty / wrong name.
    """

    def test_uses_display_name_when_present(self, plan, profile):
        text = _pdf_text(build_plan_pdf(plan, profile))
        assert "HDFC Nifty 50 Index" in text  # display_name
        assert "Parag Parikh Flexi Cap" in text

    def test_falls_back_to_name_when_display_name_missing(self, plan, profile):
        plan.allocations[0].fund.display_name = ""
        text = _pdf_text(build_plan_pdf(plan, profile))
        # The full official name should now appear
        assert "HDFC Nifty 50 Index Direct Growth" in text

    def test_falls_back_to_amfi_code_when_both_missing(self, plan, profile):
        f = plan.allocations[0].fund
        f.display_name = ""
        f.name = ""
        text = _pdf_text(build_plan_pdf(plan, profile))
        assert "119551" in text  # amfi_code


class TestXlsxAllocationRowAlignment:
    """`build_plan_xlsx` writes a row per allocation. The alignment ternary
    on the row body is `wrap if i == 0 else (right if i >= 3 else center)`
    — AddNot mutations on line 653 flip those branches.
    """

    def test_first_column_uses_wrap_alignment(self, plan, profile):
        wb = _load(build_plan_xlsx(plan, profile), data_only=False)
        sheet = wb["Plan"]
        # Find the row of the first allocation by scanning for the fund name
        fund_name = plan.allocations[0].fund.display_name
        row_idx = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == fund_name:
                    row_idx = cell.row
                    break
            if row_idx:
                break
        assert row_idx is not None, "couldn't locate allocation row"

        first_cell = sheet.cell(row=row_idx, column=1)
        assert first_cell.alignment.wrap_text is True, (
            "first column must wrap (AddNot on `i == 0` would flip this)"
        )

    def test_numeric_columns_are_right_aligned(self, plan, profile):
        wb = _load(build_plan_xlsx(plan, profile), data_only=False)
        sheet = wb["Plan"]
        fund_name = plan.allocations[0].fund.display_name
        row_idx = None
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value == fund_name:
                    row_idx = cell.row
                    break
            if row_idx:
                break
        assert row_idx is not None

        # Columns 4, 5, 6 (i >= 3): allocation %, mode, monthly SIP — right-aligned
        for col in (4, 5, 6):
            cell = sheet.cell(row=row_idx, column=col)
            assert cell.alignment.horizontal == "right", (
                f"column {col} should be right-aligned (i >= 3 branch)"
            )

"""End-to-end tests for the /api/v2 JSON API.

LLM calls are mocked; everything else (session, routing, persistence,
background tasks, cheat code) runs real.
"""

from __future__ import annotations

import asyncio
from unittest.mock import AsyncMock, patch

import pytest
from httpx import ASGITransport, AsyncClient
from pydantic_ai.usage import RunUsage

from subprime.core.models import (
    Allocation,
    InvestmentPlan,
    MutualFund,
    StrategyOutline,
)

CHEAT = "123456"


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _staged_mock(plan: "InvestmentPlan | None" = None):
    """Build a mock generate_plan_staged that fires on_partial for every stage.

    Mirrors the real staging contract so API-layer tests exercise the same
    session-state writes the production code performs.
    """
    plan_value = plan or _mock_plan()

    async def _fn(profile, strategy=None, **kwargs):
        cb = kwargs.get("on_partial")
        if cb is not None:
            await cb(plan_value, ["core"])
            await cb(plan_value, ["core", "risks"])
            await cb(plan_value, ["core", "risks", "setup"])
        return plan_value, RunUsage()

    return _fn


def _mock_strategy() -> StrategyOutline:
    return StrategyOutline(
        equity_pct=70,
        debt_pct=20,
        gold_pct=10,
        other_pct=0,
        equity_approach="Balanced mix",
        key_themes=["growth tilt", "low cost"],
        risk_return_summary="Expected 11% CAGR with moderate drawdowns",
        open_questions=[],
    )


def _mock_plan() -> InvestmentPlan:
    return InvestmentPlan(
        allocations=[
            Allocation(
                fund=MutualFund(
                    amfi_code="119551", name="UTI Nifty 50 Index Fund", category="Large Cap"
                ),
                allocation_pct=40,
                mode="sip",
                monthly_sip_inr=20000,
                rationale="Core equity exposure",
            ),
        ],
        projected_returns={"bear": 8.0, "base": 12.0, "bull": 16.0},
        rationale="Long-term wealth building plan",
        risks=["Market volatility"],
    )


@pytest.fixture(autouse=True)
def cheat_code_env(monkeypatch):
    monkeypatch.setenv("SUBPRIME_OTP_CHEAT", CHEAT)


@pytest.fixture
async def client():
    from apps.web.main import create_app

    app = create_app()
    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://test",
    ) as c:
        c._app = app  # type: ignore[attr-defined]
        yield c


async def _wait_for_plan(client, *, timeout: float = 2.0):
    """Poll the in-memory session store until the background plan task
    finishes. Replaces /plan/status polling now that the UI consumes SSE."""
    store = client._app.state.session_store  # type: ignore[attr-defined]
    deadline = asyncio.get_event_loop().time() + timeout
    while asyncio.get_event_loop().time() < deadline:
        for s in list(store._sessions.values()):
            if s.plan is not None and not s.plan_generating:
                return s
        await asyncio.sleep(0.02)
    return None


# ---------------------------------------------------------------------------
# Session lifecycle
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_get_session_creates_new_session(client):
    r = await client.get("/api/v2/session")
    assert r.status_code == 200
    data = r.json()
    assert data["current_step"] == 1
    assert data["mode"] == "basic"
    assert data["is_demo"] is False
    assert data["has_profile"] is False
    assert "benji_session" in client.cookies


@pytest.mark.asyncio
async def test_get_session_returns_existing(client):
    r1 = await client.get("/api/v2/session")
    sid = r1.json()["id"]
    r2 = await client.get("/api/v2/session")
    assert r2.json()["id"] == sid


@pytest.mark.asyncio
async def test_set_tier_basic(client):
    r = await client.post("/api/v2/session/tier", json={"mode": "basic"})
    assert r.status_code == 200
    assert r.json()["mode"] == "basic"
    assert r.json()["current_step"] == 2


@pytest.mark.asyncio
async def test_set_tier_premium(client):
    r = await client.post("/api/v2/session/tier", json={"mode": "premium"})
    assert r.json()["mode"] == "premium"


@pytest.mark.asyncio
async def test_basic_tier_drops_demo_flag(client):
    """Picking Basic after unlocking demo mode should reset is_demo.

    Otherwise a user who played with the cheat code earlier still sees the
    25-persona research bank on the 'free consumer' flow.
    """
    # Unlock demo
    await client.post(
        "/api/v2/session/otp/verify",
        json={"email": "x@y.com", "code": CHEAT},
    )
    assert (await client.get("/api/v2/session")).json()["is_demo"] is True

    # Basic tier should drop demo
    r = await client.post("/api/v2/session/tier", json={"mode": "basic"})
    assert r.json()["is_demo"] is False
    assert r.json()["mode"] == "basic"

    # Personas endpoint now returns no research bank (only archetypes)
    p = await client.get("/api/v2/personas")
    assert p.json()["personas"] is None
    assert len(p.json()["archetypes"]) == 3


@pytest.mark.asyncio
async def test_reset_generates_new_session(client):
    r1 = await client.get("/api/v2/session")
    sid1 = r1.json()["id"]
    r2 = await client.post("/api/v2/session/reset")
    assert r2.status_code == 200
    assert r2.json()["id"] != sid1


# ---------------------------------------------------------------------------
# Profile
# ---------------------------------------------------------------------------


_VALID_PROFILE = {
    "name": "Test User",
    "age": 30,
    "monthly_sip_inr": 25000,
    "existing_corpus_inr": 500000,
    "risk_appetite": "moderate",
    "investment_horizon_years": 15,
    "life_stage": "mid career",
    "financial_goals": ["Retirement"],
}


@pytest.mark.asyncio
async def test_submit_profile_advances_step(client):
    await client.post("/api/v2/session/tier", json={"mode": "basic"})
    r = await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    assert r.status_code == 200
    body = r.json()
    assert body["has_profile"] is True
    assert body["current_step"] == 3


@pytest.mark.asyncio
async def test_submit_profile_validates_age(client):
    invalid = {**_VALID_PROFILE, "age": 15}
    r = await client.post("/api/v2/session/profile", json=invalid)
    assert r.status_code == 422


@pytest.mark.asyncio
async def test_submit_profile_validates_risk_appetite(client):
    invalid = {**_VALID_PROFILE, "risk_appetite": "yolo"}
    r = await client.post("/api/v2/session/profile", json=invalid)
    assert r.status_code == 422


# ---------------------------------------------------------------------------
# Personas
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_personas_shows_archetypes_only_for_regular_user(client):
    r = await client.get("/api/v2/personas")
    assert r.status_code == 200
    data = r.json()
    assert len(data["archetypes"]) == 3
    ids = {a["id"] for a in data["archetypes"]}
    assert ids == {"early_career", "mid_career", "retired"}
    assert data["personas"] is None


@pytest.mark.asyncio
async def test_personas_unlocks_full_bank_with_cheat(client):
    r = await client.post(
        "/api/v2/session/otp/verify",
        json={"email": "foo@bar.com", "code": CHEAT},
    )
    assert r.json()["verified"] is True
    assert r.json()["is_demo"] is True

    r2 = await client.get("/api/v2/personas")
    data = r2.json()
    assert data["personas"] is not None
    assert len(data["personas"]) > 0
    assert any(p["name"] == "Tony Stark" for p in data["personas"])


@pytest.mark.asyncio
async def test_persona_select_requires_demo(client):
    # Regular (non-demo) session can't pick research personas
    r = await client.post("/api/v2/session/persona", json={"persona_id": "P01"})
    assert r.status_code == 403


@pytest.mark.asyncio
async def test_persona_select_works_after_cheat(client):
    await client.post(
        "/api/v2/session/otp/verify",
        json={"email": "foo@bar.com", "code": CHEAT},
    )
    r = await client.post("/api/v2/session/persona", json={"persona_id": "P01"})
    assert r.status_code == 200
    assert r.json()["has_profile"] is True


@pytest.mark.asyncio
async def test_persona_select_404_for_unknown(client):
    await client.post(
        "/api/v2/session/otp/verify",
        json={"email": "foo@bar.com", "code": CHEAT},
    )
    r = await client.post("/api/v2/session/persona", json={"persona_id": "NONEXISTENT"})
    assert r.status_code == 404


# ---------------------------------------------------------------------------
# OTP
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_otp_verify_rejects_wrong_code(client):
    r = await client.post(
        "/api/v2/session/otp/verify",
        json={"email": "x@y.com", "code": "999999"},
    )
    assert r.json()["verified"] is False
    assert r.json()["is_demo"] is False


@pytest.mark.asyncio
async def test_otp_verify_accepts_cheat(client):
    r = await client.post(
        "/api/v2/session/otp/verify",
        json={"email": "x@y.com", "code": CHEAT},
    )
    assert r.json()["verified"] is True


# ---------------------------------------------------------------------------
# Strategy
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_strategy_requires_profile(client):
    await client.post("/api/v2/session/tier", json={"mode": "basic"})
    r = await client.post("/api/v2/strategy/generate")
    assert r.status_code == 400


@pytest.mark.asyncio
async def test_strategy_generate(client):
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    mock = AsyncMock(return_value=(_mock_strategy(), RunUsage()))
    with patch("apps.web.api_v2.strategy.generate_strategy", new=mock):
        r = await client.post("/api/v2/strategy/generate")
    assert r.status_code == 200
    data = r.json()
    assert data["strategy"]["equity_pct"] == 70
    assert data["strategy"]["equity_approach"] == "Balanced mix"


@pytest.mark.asyncio
async def test_strategy_revise_appends_to_chat(client):
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    mock = AsyncMock(return_value=(_mock_strategy(), RunUsage()))
    with patch("apps.web.api_v2.strategy.generate_strategy", new=mock):
        await client.post("/api/v2/strategy/generate")
        r = await client.post(
            "/api/v2/strategy/revise",
            json={"feedback": "more aggressive please"},
        )
    assert r.status_code == 200
    chat = r.json()["chat"]
    assert len(chat) == 2  # user + advisor
    assert chat[0]["role"] == "user"
    assert chat[0]["content"] == "more aggressive please"


@pytest.mark.asyncio
async def test_strategy_answer_questions_does_not_pollute_chat(client):
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    mock = AsyncMock(return_value=(_mock_strategy(), RunUsage()))
    with patch("apps.web.api_v2.strategy.generate_strategy", new=mock):
        await client.post("/api/v2/strategy/generate")
        r = await client.post(
            "/api/v2/strategy/answer-questions",
            json={"feedback": "horizon is actually 20 years"},
        )
    assert r.json()["chat"] == []


# ---------------------------------------------------------------------------
# Plan — background task + poll
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_plan_generate_returns_202_and_sets_flag(client):
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    with patch("apps.web.api_v2.plan.generate_plan_staged", new=_staged_mock()):
        r = await client.post("/api/v2/plan/generate")
        assert r.status_code == 202
        s = await _wait_for_plan(client)
    assert s is not None
    assert s.plan is not None
    assert s.plan_generating is False


@pytest.mark.asyncio
async def test_plan_get_404_before_ready(client):
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    r = await client.get("/api/v2/plan")
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_plan_stream_reports_staged_progress(client):
    """Stages arrive incrementally — the SSE stream should expose each one."""
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    with patch("apps.web.api_v2.plan.generate_plan_staged", new=_staged_mock()):
        await client.post("/api/v2/plan/generate")
        s = await _wait_for_plan(client)
    assert s is not None
    assert set(s.plan_stages) >= {"core", "risks", "setup"}


@pytest.mark.asyncio
async def test_plan_get_returns_full_plan_after_ready(client):
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    with patch("apps.web.api_v2.plan.generate_plan_staged", new=_staged_mock()):
        await client.post("/api/v2/plan/generate")
        await _wait_for_plan(client)
        r = await client.get("/api/v2/plan")
    assert r.status_code == 200
    data = r.json()
    assert len(data["plan"]["allocations"]) == 1
    assert data["plan"]["allocations"][0]["fund"]["name"] == "UTI Nifty 50 Index Fund"
    assert data["profile"]["name"] == "Test User"


@pytest.mark.asyncio
async def test_plan_generate_requires_profile(client):
    r = await client.post("/api/v2/plan/generate")
    assert r.status_code == 400


@pytest.mark.asyncio
async def test_plan_error_surfaces_in_session(client):
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    broken = AsyncMock(side_effect=RuntimeError("model hiccup"))
    with patch("apps.web.api_v2.plan.generate_plan_staged", new=broken):
        await client.post("/api/v2/plan/generate")
        # Wait for the error path to flip plan_generating off.
        store = client._app.state.session_store
        for _ in range(50):
            await asyncio.sleep(0.02)
            sess = next(iter(store._sessions.values()), None)
            if sess and not sess.plan_generating:
                break
    assert sess is not None
    assert sess.plan is None
    assert "model hiccup" in (sess.plan_error or "")


# ---------------------------------------------------------------------------
# Full flow — tier → profile → strategy → plan
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_full_wizard_flow_end_to_end(client):
    # 1. Tier
    r = await client.post("/api/v2/session/tier", json={"mode": "basic"})
    assert r.json()["current_step"] == 2

    # 2. Custom profile
    r = await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    assert r.json()["has_profile"] is True
    assert r.json()["current_step"] == 3

    # 3. Strategy
    with patch(
        "apps.web.api_v2.strategy.generate_strategy",
        new=AsyncMock(return_value=(_mock_strategy(), RunUsage())),
    ):
        r = await client.post("/api/v2/strategy/generate")
    assert r.status_code == 200

    session = (await client.get("/api/v2/session")).json()
    assert session["has_strategy"] is True

    # 4. Plan
    with patch(
        "apps.web.api_v2.plan.generate_plan_staged",
        new=_staged_mock(),
    ):
        r = await client.post("/api/v2/plan/generate")
        assert r.status_code == 202
        await _wait_for_plan(client)

    plan_resp = await client.get("/api/v2/plan")
    assert plan_resp.status_code == 200
    assert len(plan_resp.json()["plan"]["allocations"]) > 0


# ---------------------------------------------------------------------------
# Plan — download (PDF / Excel)
# ---------------------------------------------------------------------------


async def _seed_plan(client):
    """Helper: push a profile + generated plan into the session so the
    download endpoints have something to serve."""
    await client.post("/api/v2/session/profile", json=_VALID_PROFILE)
    with patch("apps.web.api_v2.plan.generate_plan_staged", new=_staged_mock()):
        await client.post("/api/v2/plan/generate")
        await _wait_for_plan(client)


@pytest.mark.asyncio
async def test_download_pdf_404_without_plan(client):
    r = await client.get("/api/v2/plan/download.pdf")
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_download_xlsx_404_without_plan(client):
    r = await client.get("/api/v2/plan/download.xlsx")
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_download_pdf_returns_branded_pdf(client):
    await _seed_plan(client)
    r = await client.get("/api/v2/plan/download.pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    cd = r.headers["content-disposition"]
    assert cd.startswith("attachment;")
    assert 'filename="benji-plan-' in cd
    assert cd.endswith('.pdf"')
    body = r.content
    assert body[:5] == b"%PDF-"
    # Extract rendered text via pypdf — ReportLab compresses content streams
    # so raw-byte grep is unreliable.
    from io import BytesIO
    from pypdf import PdfReader

    reader = PdfReader(BytesIO(body))
    text = "\n".join(page.extract_text() or "" for page in reader.pages)
    assert "Benji" in text
    assert "Test User" in text
    assert "research/educational purposes" in text


@pytest.mark.asyncio
async def test_download_xlsx_returns_openable_workbook(client):
    from io import BytesIO
    from openpyxl import load_workbook

    await _seed_plan(client)
    r = await client.get("/api/v2/plan/download.xlsx")
    assert r.status_code == 200
    assert r.headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    cd = r.headers["content-disposition"]
    assert 'filename="benji-plan-' in cd and cd.endswith('.xlsx"')
    wb = load_workbook(filename=BytesIO(r.content))
    # New consolidated shape: Plan + Explore
    assert wb.sheetnames == ["Plan", "Explore"]
    flat = [str(c.value) for row in wb["Plan"].iter_rows() for c in row if c.value is not None]
    # Mocked plan fund name is 'UTI Nifty 50 Index Fund'
    assert any("UTI Nifty 50" in s for s in flat)
    # Allocation percentage (40) should appear as a number somewhere
    numbers = [
        c.value for row in wb["Plan"].iter_rows() for c in row if isinstance(c.value, (int, float))
    ]
    assert 40 in numbers or 40.0 in numbers

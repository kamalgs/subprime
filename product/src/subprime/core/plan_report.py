"""Generate downloadable plan reports — PDF + Excel.

PDF: single-page (usually) branded A4. Header with the Benji wordmark +
short disclaimer above a ** footnote, then an investor summary, an
allocations table with wrapping fund names, a projections section that
combines a CAGR table with a line-chart of the projected corpus, then
setup/checkpoints/rebalancing/rationale/risks. Full disclaimer at the
bottom.

Excel: two sheets.
  - Plan — headline numbers, allocations table, a corpus projection
    chart over time, short disclaimer.
  - Explore — an editable model. User can tweak the monthly SIP, the
    horizon, and the three CAGR assumptions; the final corpus cells
    recompute via Excel formulas. An embedded chart updates live.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO

from reportlab.graphics.charts.lineplots import LinePlot
from reportlab.graphics.shapes import Drawing, String
from reportlab.graphics.widgets.markers import makeMarker
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from subprime.core.models import InvestmentPlan, InvestorProfile

# Brand palette (matches tailwind primary-600 / red-600 / slate).
BRAND = colors.HexColor("#dc2626")
PRIMARY = colors.HexColor("#2563eb")
PRIMARY_DARK = colors.HexColor("#1e40af")
MUTED = colors.HexColor("#64748b")
DARK = colors.HexColor("#0f172a")
LIGHT = colors.HexColor("#f1f5f9")

# Short disclaimer shown UP FRONT with a ** footnote.
SHORT_DISCLAIMER = "<b>For research / educational use only — not certified financial advice.**</b>"


# ── ReportLab styles ──────────────────────────────────────────────────────────


def _styles() -> dict[str, ParagraphStyle]:
    ss = getSampleStyleSheet()
    return {
        "wordmark": ParagraphStyle(
            "wordmark",
            parent=ss["Title"],
            fontName="Helvetica-Bold",
            fontSize=26,
            leading=28,
            textColor=BRAND,
            spaceAfter=2,
        ),
        "short_disclaimer": ParagraphStyle(
            "short_disclaimer",
            parent=ss["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=BRAND,
            alignment=0,
            spaceAfter=10,
        ),
        "subtitle": ParagraphStyle(
            "subtitle",
            parent=ss["BodyText"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=MUTED,
            spaceAfter=10,
        ),
        "section": ParagraphStyle(
            "section",
            parent=ss["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            leading=16,
            textColor=BRAND,
            spaceBefore=10,
            spaceAfter=5,
        ),
        "body": ParagraphStyle(
            "body",
            parent=ss["BodyText"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=12.5,
            textColor=DARK,
            spaceAfter=3,
        ),
        "cell": ParagraphStyle(
            "cell",
            parent=ss["BodyText"],
            fontName="Helvetica",
            fontSize=8.5,
            leading=11,
            textColor=DARK,
            spaceAfter=0,
        ),
        "cell_name": ParagraphStyle(
            "cell_name",
            parent=ss["BodyText"],
            fontName="Helvetica-Bold",
            fontSize=8.5,
            leading=11,
            textColor=DARK,
            spaceAfter=0,
        ),
        "bullet": ParagraphStyle(
            "bullet",
            parent=ss["BodyText"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=13,
            leftIndent=14,
            bulletIndent=2,
            textColor=DARK,
            spaceAfter=2,
        ),
        "footnote": ParagraphStyle(
            "footnote",
            parent=ss["BodyText"],
            fontName="Helvetica-Oblique",
            fontSize=8,
            leading=11,
            textColor=MUTED,
            alignment=0,
            spaceBefore=8,
        ),
        "disclaimer": ParagraphStyle(
            "disclaimer",
            parent=ss["BodyText"],
            fontName="Helvetica-Oblique",
            fontSize=8.5,
            leading=12,
            textColor=BRAND,
            alignment=1,
            spaceBefore=10,
        ),
    }


def _header_band(canvas, doc) -> None:
    """Red accent band up top + 'Benji' footer with page numbers."""
    canvas.saveState()
    canvas.setFillColor(BRAND)
    canvas.rect(0, doc.pagesize[1] - 0.35 * cm, doc.pagesize[0], 0.35 * cm, stroke=0, fill=1)
    canvas.setFillColor(MUTED)
    canvas.setFont("Helvetica", 8)
    canvas.drawString(1.5 * cm, 1.1 * cm, "Benji — AI financial advisor")
    canvas.drawRightString(doc.pagesize[0] - 1.5 * cm, 1.1 * cm, f"Page {canvas.getPageNumber()}")
    canvas.restoreState()


# ── Formatting helpers ────────────────────────────────────────────────────────


def _fmt_money_inr(amount: float) -> str:
    """Indian notation with lakhs / crores."""
    if amount >= 1_00_00_000:
        return f"\u20b9{amount / 1_00_00_000:.2f} Cr"
    if amount >= 1_00_000:
        return f"\u20b9{amount / 1_00_000:.2f} L"
    return f"\u20b9{amount:,.0f}"


def _project_corpus(monthly_sip: float, horizon_years: int, annual_pct: float) -> float:
    """FV of a monthly SIP at a given annual % (compounded monthly).

    FV = P * [((1 + r)^n − 1) / r] * (1 + r)
    where P = monthly contribution, r = monthly rate, n = months.
    """
    if not monthly_sip or not horizon_years or annual_pct is None:
        return 0.0
    r = annual_pct / 100 / 12
    n = horizon_years * 12
    if r == 0:
        return monthly_sip * n
    return monthly_sip * ((pow(1 + r, n) - 1) / r) * (1 + r)


def _projection_trace(
    monthly_sip: float, horizon_years: int, annual_pct: float
) -> list[tuple[float, float]]:
    """Per-year (year, corpus) pairs for the line chart."""
    trace: list[tuple[float, float]] = [(0.0, 0.0)]
    for year in range(1, horizon_years + 1):
        trace.append((float(year), _project_corpus(monthly_sip, year, annual_pct)))
    return trace


# ── PDF section builders ──────────────────────────────────────────────────────


def _allocations_table(plan: InvestmentPlan, styles: dict) -> Table:
    """Wrapping fund-name column, right-aligned numbers, alt-row shading."""
    header = [
        Paragraph("<b>Fund</b>", styles["cell_name"]),
        Paragraph("<b>AMC</b>", styles["cell_name"]),
        Paragraph("<b>Category</b>", styles["cell_name"]),
        Paragraph("<b>Alloc</b>", styles["cell_name"]),
        Paragraph("<b>Mode</b>", styles["cell_name"]),
        Paragraph("<b>Monthly</b>", styles["cell_name"]),
    ]
    rows = [header]
    for a in plan.allocations:
        fund = a.fund
        name = fund.display_name or fund.name or fund.amfi_code
        sip = _fmt_money_inr(a.monthly_sip_inr) if a.monthly_sip_inr else "—"
        rows.append(
            [
                Paragraph(name, styles["cell_name"]),
                Paragraph(fund.fund_house or "—", styles["cell"]),
                Paragraph(fund.category or "—", styles["cell"]),
                Paragraph(f"{a.allocation_pct:.0f}%", styles["cell"]),
                Paragraph((a.mode or "—").upper(), styles["cell"]),
                Paragraph(sip, styles["cell"]),
            ]
        )
    # A4 usable width ≈ 18 cm with 1.5 cm side margins. These widths sum to
    # 17.6 cm so there's a touch of breathing room.
    tbl = Table(
        rows,
        colWidths=[5.8 * cm, 3.0 * cm, 2.8 * cm, 1.4 * cm, 1.6 * cm, 3.0 * cm],
        repeatRows=1,
    )
    tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), PRIMARY),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("ALIGN", (3, 0), (5, -1), "RIGHT"),
                ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.white),
            ]
        )
    )
    return tbl


def _projection_chart(monthly_sip: float, horizon_years: int, returns: dict[str, float]) -> Drawing:
    """Line plot of bear / base / bull corpus over the horizon."""
    bear_r = returns.get("bear", 8.0)
    base_r = returns.get("base", 12.0)
    bull_r = returns.get("bull", 16.0)

    bear = _projection_trace(monthly_sip, horizon_years, bear_r)
    base = _projection_trace(monthly_sip, horizon_years, base_r)
    bull = _projection_trace(monthly_sip, horizon_years, bull_r)

    d = Drawing(480, 180)
    lp = LinePlot()
    lp.x, lp.y = 40, 28
    lp.width, lp.height = 420, 130

    # Convert corpus values to lakhs on the y-axis so labels stay short.
    def to_lakhs(pairs):
        return [(x, y / 1_00_000) for x, y in pairs]

    lp.data = [to_lakhs(bear), to_lakhs(base), to_lakhs(bull)]

    # Colours + markers
    lp.lines[0].strokeColor = colors.HexColor("#ef4444")  # red-500
    lp.lines[1].strokeColor = colors.HexColor("#f59e0b")  # amber-500
    lp.lines[2].strokeColor = colors.HexColor("#10b981")  # green-500
    for i in range(3):
        lp.lines[i].strokeWidth = 1.6
        lp.lines[i].symbol = makeMarker("Circle")
        lp.lines[i].symbol.size = 2.8
        lp.lines[i].symbol.fillColor = lp.lines[i].strokeColor
        lp.lines[i].symbol.strokeColor = lp.lines[i].strokeColor

    lp.xValueAxis.valueMin = 0
    lp.xValueAxis.valueMax = horizon_years
    lp.xValueAxis.valueStep = max(1, horizon_years // 5)
    lp.xValueAxis.labels.fontSize = 8
    lp.xValueAxis.labels.fillColor = MUTED
    lp.xValueAxis.strokeColor = MUTED

    max_y = max(bull[-1][1] / 1_00_000, 1)
    lp.yValueAxis.valueMin = 0
    lp.yValueAxis.valueMax = max_y * 1.08
    # Pick a nice step (in lakhs)
    step = max(1, int(max_y / 4))
    lp.yValueAxis.valueStep = step
    lp.yValueAxis.labels.fontSize = 8
    lp.yValueAxis.labels.fillColor = MUTED
    lp.yValueAxis.strokeColor = MUTED

    d.add(lp)

    # Axis titles
    d.add(
        String(
            240, 8, "Years", fontName="Helvetica", fontSize=8, fillColor=MUTED, textAnchor="middle"
        )
    )
    d.add(
        String(
            6,
            168,
            "Corpus (\u20b9 lakh)",
            fontName="Helvetica",
            fontSize=8,
            fillColor=MUTED,
            textAnchor="start",
        )
    )

    # Legend (manual — reportlab auto-legend is clunky)
    legend_y = 170
    for i, (label, colour) in enumerate(
        [
            ("Bear", colors.HexColor("#ef4444")),
            ("Base", colors.HexColor("#f59e0b")),
            ("Bull", colors.HexColor("#10b981")),
        ]
    ):
        x = 380 - (2 - i) * 40
        d.add(String(x + 10, legend_y, label, fontName="Helvetica", fontSize=8, fillColor=DARK))
        from reportlab.graphics.shapes import Line

        d.add(Line(x, legend_y + 3, x + 8, legend_y + 3, strokeColor=colour, strokeWidth=2))
    return d


def _projections_block(plan: InvestmentPlan, profile: InvestorProfile, styles: dict) -> list:
    """CAGR table on the left, absolute-corpus table on the right, chart
    below. Returns a list of flowables ready to extend into the story."""
    pr = plan.projected_returns or {}
    if not pr:
        return []
    horizon = int(getattr(profile, "investment_horizon_years", 0) or 0)
    monthly_sip = sum(a.monthly_sip_inr or 0 for a in plan.allocations)
    # Side-by-side mini-tables
    cagr_data = [["Scenario", "CAGR"]]
    corpus_data = [["Scenario", f"Corpus @ {horizon} yrs"]] if horizon else [["Scenario", "Corpus"]]
    for key in ("bear", "base", "bull"):
        pct = pr.get(key)
        if pct is None:
            continue
        cagr_data.append([key.title(), f"{pct:.1f}%"])
        if horizon and monthly_sip:
            corpus_data.append(
                [
                    key.title(),
                    _fmt_money_inr(_project_corpus(monthly_sip, horizon, pct)),
                ]
            )

    def _mini(data, header_color):
        t = Table(data, colWidths=[2.5 * cm, 4.0 * cm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), header_color),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [LIGHT, colors.white]),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        return t

    flowables: list = []
    if horizon and monthly_sip:
        side = Table(
            [[_mini(cagr_data, PRIMARY), _mini(corpus_data, PRIMARY_DARK)]],
            colWidths=[7.5 * cm, 7.5 * cm],
        )
        side.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
        flowables.append(side)
        flowables.append(Spacer(1, 8))
        flowables.append(_projection_chart(monthly_sip, horizon, pr))
    else:
        flowables.append(_mini(cagr_data, PRIMARY))
    return flowables


def _split_bullets(text: str) -> list[str]:
    if not text:
        return []
    import re

    out: list[str] = []
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        for prefix in ("- ", "* ", "+ ", "\u2022 "):
            if s.startswith(prefix):
                s = s[len(prefix) :]
                break
        s = re.sub(r"^\d+[.)]\s*", "", s)
        if s:
            out.append(s)
    return out or [text.strip()]


# ── PDF entry point ───────────────────────────────────────────────────────────


def build_plan_pdf(plan: InvestmentPlan, profile: InvestorProfile) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.4 * cm,
        bottomMargin=2 * cm,
        title=f"Benji plan — {profile.name}",
        author="Benji",
    )
    styles = _styles()
    story: list = []

    # Header: wordmark + short disclaimer
    story.append(Paragraph("Benji", styles["wordmark"]))
    story.append(Paragraph(SHORT_DISCLAIMER, styles["short_disclaimer"]))

    story.append(
        Paragraph(
            f"Investment plan for <b>{profile.name}</b> \u00b7 "
            f"{getattr(profile, 'life_stage', 'investor')} \u00b7 "
            f"generated {datetime.utcnow().strftime('%d %b %Y')}",
            styles["subtitle"],
        )
    )

    # Investor summary
    monthly = getattr(profile, "monthly_investible_surplus_inr", None)
    horizon = getattr(profile, "investment_horizon_years", None)
    risk = getattr(profile, "risk_appetite", None)
    summary_bits = []
    if monthly:
        summary_bits.append(f"Investible surplus: <b>{_fmt_money_inr(float(monthly))}/mo</b>")
    if horizon:
        summary_bits.append(f"Horizon: <b>{horizon} years</b>")
    if risk:
        summary_bits.append(f"Risk: <b>{risk}</b>")
    if summary_bits:
        story.append(Paragraph(" \u00b7 ".join(summary_bits), styles["body"]))
        story.append(Spacer(1, 4))

    # Allocations
    story.append(Paragraph("Allocations", styles["section"]))
    story.append(_allocations_table(plan, styles))

    # Projections — tables + chart
    proj = _projections_block(plan, profile, styles)
    if proj:
        story.append(Paragraph("Projected returns", styles["section"]))
        story.extend(proj)

    # Setup
    if plan.setup_phase:
        story.append(Paragraph("Setup", styles["section"]))
        for line in _split_bullets(plan.setup_phase):
            story.append(Paragraph(line, styles["bullet"], bulletText="\u203a"))

    # Checkpoints
    if plan.review_checkpoints:
        story.append(Paragraph("Review checkpoints", styles["section"]))
        for cp in plan.review_checkpoints:
            story.append(Paragraph(cp, styles["bullet"], bulletText="\u203a"))

    # Rebalancing
    if plan.rebalancing_guidelines:
        story.append(Paragraph("Rebalancing", styles["section"]))
        for line in _split_bullets(plan.rebalancing_guidelines):
            story.append(Paragraph(line, styles["bullet"], bulletText="\u203a"))

    # Rationale
    if plan.rationale:
        story.append(Paragraph("Why this plan", styles["section"]))
        for line in _split_bullets(plan.rationale):
            story.append(Paragraph(line, styles["bullet"], bulletText="\u203a"))

    # Risks
    if plan.risks:
        story.append(Paragraph("Risks to consider", styles["section"]))
        for r in plan.risks:
            story.append(Paragraph(r, styles["bullet"], bulletText="\u26a0"))

    # ** footnote + full disclaimer at the bottom
    story.append(Spacer(1, 10))
    story.append(
        Paragraph(
            "** "
            + plan.disclaimer
            + " Projections assume monthly SIP at the given CAGR, compounded "
            "monthly — no guarantee of future returns. Consult a SEBI-registered "
            "investment advisor before acting.",
            styles["footnote"],
        )
    )

    doc.build(story, onFirstPage=_header_band, onLaterPages=_header_band)
    return buf.getvalue()


# ── Excel ─────────────────────────────────────────────────────────────────────


def build_plan_xlsx(plan: InvestmentPlan, profile: InvestorProfile) -> bytes:
    """Two-sheet workbook:

    Plan    — headline KPIs, allocations table, projection chart.
    Explore — editable model. Change monthly SIP / horizon / CAGRs and
              the final corpus + chart recompute via Excel formulas.
    """
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.chart.shapes import GraphicalProperties
    from openpyxl.drawing.line import LineProperties
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )

    wb = Workbook()
    wb.remove(wb.active)

    BRAND_HEX = "DC2626"
    PRIMARY_HEX = "2563EB"
    LIGHT_HEX = "F1F5F9"
    MUTED_HEX = "64748B"

    hdr_fill = PatternFill("solid", fgColor=PRIMARY_HEX)
    hdr_font = Font(bold=True, color="FFFFFF")
    brand_font = Font(bold=True, color=BRAND_HEX, size=18)
    italic_brand = Font(italic=True, color=BRAND_HEX, size=9)
    title_font = Font(bold=True, size=12, color="0F172A")
    muted = Font(color=MUTED_HEX, size=10)
    thin = Side(style="thin", color="CBD5E1")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    monthly_sip = sum(a.monthly_sip_inr or 0 for a in plan.allocations)
    horizon = int(getattr(profile, "investment_horizon_years", 0) or 0)
    pr = plan.projected_returns or {}
    bear_r = pr.get("bear", 8.0)
    base_r = pr.get("base", 12.0)
    bull_r = pr.get("bull", 16.0)

    # ═══ Sheet 1: Plan ══════════════════════════════════════════════
    p = wb.create_sheet("Plan")
    p.column_dimensions["A"].width = 42
    p.column_dimensions["B"].width = 20
    p.column_dimensions["C"].width = 18
    p.column_dimensions["D"].width = 14
    p.column_dimensions["E"].width = 16
    p.column_dimensions["F"].width = 18

    # Wordmark + short disclaimer
    p["A1"] = "Benji"
    p["A1"].font = brand_font
    p["A2"] = "Investment plan — for research / educational use only**"
    p["A2"].font = italic_brand

    # Investor summary
    p["A4"] = "Investor"
    p["A4"].font = title_font
    p["B4"] = profile.name
    p["C4"] = "Generated"
    p["C4"].font = title_font
    p["D4"] = datetime.utcnow().strftime("%d %b %Y")

    row = 6
    summary_rows = [
        ("Life stage", getattr(profile, "life_stage", "") or ""),
        ("Age", getattr(profile, "age", "") or ""),
        ("Risk appetite", getattr(profile, "risk_appetite", "") or ""),
        ("Monthly surplus (₹)", getattr(profile, "monthly_investible_surplus_inr", "") or ""),
        ("Horizon (years)", horizon),
        ("Funds", len(plan.allocations)),
        ("Total monthly SIP (₹)", monthly_sip),
    ]
    for label, val in summary_rows:
        p.cell(row=row, column=1, value=label).font = Font(bold=True)
        p.cell(row=row, column=2, value=val)
        row += 1
    row += 1

    # Allocations table
    p.cell(row=row, column=1, value="Allocations").font = title_font
    row += 1
    alloc_header = ["Fund", "AMC", "Category", "% Alloc", "Mode", "Monthly SIP (₹)"]
    for i, h in enumerate(alloc_header):
        c = p.cell(row=row, column=i + 1, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = box
    row += 1
    alloc_start = row
    for a in plan.allocations:
        f = a.fund
        cells = [
            (f.display_name or f.name or f.amfi_code),
            f.fund_house or "",
            f.category or "",
            round(a.allocation_pct, 2),
            (a.mode or "").upper(),
            a.monthly_sip_inr or 0,
        ]
        for i, v in enumerate(cells):
            c = p.cell(row=row, column=i + 1, value=v)
            c.border = box
            c.alignment = wrap if i == 0 else (right if i >= 3 else Alignment(vertical="center"))
        if row % 2 == alloc_start % 2:
            for i in range(len(cells)):
                p.cell(row=row, column=i + 1).fill = PatternFill("solid", fgColor=LIGHT_HEX)
        row += 1
    row += 1

    # Projections
    if pr and horizon and monthly_sip:
        p.cell(row=row, column=1, value="Projected returns").font = title_font
        row += 1
        headers = ["Scenario", "CAGR", f"Final corpus after {horizon} years"]
        for i, h in enumerate(headers):
            c = p.cell(row=row, column=i + 1, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = center
            c.border = box
        row += 1
        for label, rate in [("Bear", bear_r), ("Base", base_r), ("Bull", bull_r)]:
            corpus = _project_corpus(monthly_sip, horizon, rate)
            for i, v in enumerate([label, rate / 100, corpus]):
                c = p.cell(row=row, column=i + 1, value=v)
                c.border = box
                if i == 1:
                    c.number_format = "0.0%"
                    c.alignment = right
                elif i == 2:
                    c.number_format = '"₹"#,##0'
                    c.alignment = right
            row += 1
        row += 1

        # Growth-over-time chart — compute per-year trace on a hidden part
        # of the sheet, then reference it in a LineChart.
        start_chart_row = row
        chart_header = ["Year", "Bear (₹)", "Base (₹)", "Bull (₹)"]
        for i, h in enumerate(chart_header):
            c = p.cell(row=row, column=i + 1, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = center
        row += 1
        data_start = row
        for year in range(0, horizon + 1):
            p.cell(row=row, column=1, value=year)
            p.cell(
                row=row, column=2, value=_project_corpus(monthly_sip, year, bear_r)
            ).number_format = '"₹"#,##0'
            p.cell(
                row=row, column=3, value=_project_corpus(monthly_sip, year, base_r)
            ).number_format = '"₹"#,##0'
            p.cell(
                row=row, column=4, value=_project_corpus(monthly_sip, year, bull_r)
            ).number_format = '"₹"#,##0'
            row += 1
        data_end = row - 1

        chart = LineChart()
        chart.title = "Projected corpus"
        chart.style = 10
        chart.y_axis.title = "Corpus (₹)"
        chart.x_axis.title = "Year"
        chart.height = 8
        chart.width = 18
        data = Reference(p, min_col=2, min_row=start_chart_row, max_col=4, max_row=data_end)
        cats = Reference(p, min_col=1, min_row=data_start, max_row=data_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        for i, hex_ in enumerate(["EF4444", "F59E0B", "10B981"]):
            s = chart.series[i]
            s.graphicalProperties = GraphicalProperties(solidFill=hex_)
            s.graphicalProperties.line = LineProperties(solidFill=hex_, w=22000)
        # Place chart to the right of the data
        p.add_chart(chart, f"F{start_chart_row}")

    # Disclaimer at bottom
    row = max(row, 40) + 2
    p.cell(
        row=row,
        column=1,
        value=(
            "** " + plan.disclaimer + " Projections assume monthly SIP at the given CAGR, "
            "compounded monthly. Not guaranteed."
        ),
    ).font = italic_brand
    p.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    p.row_dimensions[row].height = 40
    p.cell(row=row, column=1).alignment = wrap

    # ═══ Sheet 2: Explore — interactive model ═════════════════════
    ex = wb.create_sheet("Explore")
    for col, w in zip("ABCDEFGH", [30, 14, 14, 14, 14, 14, 14, 14]):
        ex.column_dimensions[col].width = w
    ex["A1"] = "Benji — Explore your plan"
    ex["A1"].font = brand_font
    ex["A2"] = "Edit the blue cells below; the corpus and chart recompute automatically."
    ex["A2"].font = muted

    ex["A4"] = "Inputs"
    ex["A4"].font = title_font

    input_rows = [
        ("Monthly SIP (₹)", monthly_sip, '"₹"#,##0'),
        ("Horizon (years)", horizon, "0"),
        ("Bear CAGR", bear_r / 100, "0.0%"),
        ("Base CAGR", base_r / 100, "0.0%"),
        ("Bull CAGR", bull_r / 100, "0.0%"),
    ]
    input_cells: dict[str, str] = {}
    keys = ["sip", "horizon", "bear_r", "base_r", "bull_r"]
    for i, ((label, val, fmt), key) in enumerate(zip(input_rows, keys)):
        r = 5 + i
        lbl = ex.cell(row=r, column=1, value=label)
        lbl.font = Font(bold=True)
        c = ex.cell(row=r, column=2, value=val)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor="DBEAFE")  # light blue — editable
        c.font = Font(bold=True, color="1E40AF")
        c.border = box
        input_cells[key] = c.coordinate

    # Outputs — final corpus per scenario
    ex["D4"] = "Final corpus"
    ex["D4"].font = title_font
    ex["D5"] = "Bear"
    ex["D6"] = "Base"
    ex["D7"] = "Bull"

    # FV of growing series: P * (((1+r)^n − 1) / r) * (1 + r)
    # where P = SIP, r = monthly rate = CAGR/12, n = months = years*12
    def fv_formula(rate_cell: str) -> str:
        sip = input_cells["sip"]
        h = input_cells["horizon"]
        # Parentheses-heavy but correct
        return (
            f"=IF(OR({rate_cell}=0,{h}=0),"
            f"{sip}*{h}*12,"
            f"{sip} * ((POWER(1+{rate_cell}/12, {h}*12) - 1) / ({rate_cell}/12)) "
            f"* (1 + {rate_cell}/12))"
        )

    ex["E5"] = fv_formula(input_cells["bear_r"])
    ex["E6"] = fv_formula(input_cells["base_r"])
    ex["E7"] = fv_formula(input_cells["bull_r"])
    for cell in ("E5", "E6", "E7"):
        ex[cell].number_format = '"₹"#,##0'
        ex[cell].font = Font(bold=True, color="0F172A")
        ex[cell].border = box

    # Per-year growth table referenced by the chart (formula-driven so it
    # updates when the user edits inputs).
    ex["A11"] = "Corpus over time (computed live)"
    ex["A11"].font = title_font
    ex["A12"] = "Year"
    ex["B12"] = "Bear"
    ex["C12"] = "Base"
    ex["D12"] = "Bull"
    for i, h in enumerate(["Year", "Bear", "Base", "Bull"]):
        c = ex.cell(row=12, column=i + 1)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center

    # 41 rows so up to 40-year horizons render; cells beyond horizon return 0
    MAX_YEARS = 40
    sip = input_cells["sip"]
    h = input_cells["horizon"]
    bearR = input_cells["bear_r"]
    baseR = input_cells["base_r"]
    bullR = input_cells["bull_r"]
    for i in range(MAX_YEARS + 1):
        year_cell = 13 + i
        ex.cell(row=year_cell, column=1, value=i)
        for col, rate in zip(("B", "C", "D"), (bearR, baseR, bullR)):
            formula = (
                f"=IF({i}>{h},NA(),"
                f"IF(OR({rate}=0,{i}=0),{sip}*{i}*12,"
                f"{sip}*((POWER(1+{rate}/12, {i}*12)-1)/({rate}/12))*(1+{rate}/12)))"
            )
            cell = ex.cell(row=year_cell, column="ABCD".index(col) + 1, value=formula)
            cell.number_format = '"₹"#,##0'

    # Chart
    chart = LineChart()
    chart.title = "Projected corpus — edit inputs to explore"
    chart.style = 10
    chart.y_axis.title = "Corpus (₹)"
    chart.x_axis.title = "Year"
    chart.height = 10
    chart.width = 18
    data = Reference(ex, min_col=2, min_row=12, max_col=4, max_row=13 + MAX_YEARS)
    cats = Reference(ex, min_col=1, min_row=13, max_row=13 + MAX_YEARS)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    for i, hex_ in enumerate(["EF4444", "F59E0B", "10B981"]):
        s = chart.series[i]
        s.graphicalProperties = GraphicalProperties(solidFill=hex_)
        s.graphicalProperties.line = LineProperties(solidFill=hex_, w=22000)
    chart.anchor = "F4"
    ex.add_chart(chart)

    # Disclaimer at bottom of Explore
    ex["A55"] = (
        "** " + plan.disclaimer + " Projections are formula-based estimates at the rates shown, "
        "not guaranteed returns."
    )
    ex["A55"].font = italic_brand
    ex.merge_cells("A55:H55")
    ex["A55"].alignment = wrap
    ex.row_dimensions[55].height = 40

    # Make Plan the active tab
    wb.active = wb.sheetnames.index("Plan")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
